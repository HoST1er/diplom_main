import os
import fnmatch
import openpyxl
from django.core.management.base import BaseCommand
from main.models import FirstVariantBd



class Command(BaseCommand):
    help = 'Парсинг файлов Excel и загрузка данных в базу'

    def handle(self, *args, **kwargs):
        rootPath = 'C:/Users/andru/PycharmProjects/diplom/УЧЕБНЫЕ ПЛАНЫ'
        pattern = '*.xlsx'

        for root, dirs, files in os.walk(rootPath):
            for filename in fnmatch.filter(files, pattern):
                file_path = os.path.join(root, filename)
                try:
                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb['Дисциплины']

                    try:
                        cell_value = sheet['A11'].value
                        cell_value = cell_value.split('Профиль: ')
                        profile = cell_value[1].replace('"', '')
                    except:
                        profile = ''

                    columns = ['B', 'C', 'D', 'F', 'E', 'J',
                               'G', 'H', 'I', 'K', 'L', 'M',
                               'N', 'Q', 'R', 'S', 'T', 'O', 'P']
                    column_indices = [openpyxl.utils.column_index_from_string(col) for col in columns]
                    start_row = 23
                    data = []

                    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
                        row_data = []
                        is_bright = True

                        for idx in column_indices:
                            cell = row[idx - 1]
                            cell_color = cell.fill.start_color

                            if cell_color.index == '00000000' or cell_color.index == 'FFD5EFFF':
                                row_data.append(cell.value)
                            else:
                                is_bright = False
                                break

                        if is_bright:
                            data.append(row_data)

                    for row in data:
                        try:
                            FirstVariantBd.objects.create(
                                name_object=row[0],
                                department=row[1].replace('\n', '') if row[1] else '',
                                competentions=row[2].replace('\n', '') if row[2] else '',
                                profile=profile,
                                test_obj=row[3] if row[3] else '',
                                exam=row[4] if row[4] else '',
                                control_work=row[5] if row[5] else '',
                                test_obj_with_mark=row[6] if row[6] else '',
                                course_work=row[7] if row[7] else '',
                                course_project=row[8] if row[8] else '',
                                essay=row[9] if row[9] else '',
                                calcul_analytic_work=row[10] if row[10] else '',
                                creative_homework=row[11] if row[11] else '',
                                project_work=row[12] if row[12] else '',
                                classroom_hours=int(row[13]) if row[13] else None,
                                lectures=int(row[14]) if row[14] else None,
                                seminars=int(row[15]) if row[15] else None,
                                independent_work=int(row[16]) if row[16] else None,
                                ECTS=float(row[17]) if row[17] else None,
                                total_hours=int(row[18]) if row[18] else None,
                            )
                        except:
                            continue

                    self.stdout.write(self.style.SUCCESS(f'Файл {filename} обработан!'))
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Ошибка в файле {filename}: {e}'))
